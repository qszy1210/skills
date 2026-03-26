#!/usr/bin/env python3
"""Excel 表结构检测器 — 第一层确定性扫描。

识别表头、数据区、表尾/小计等结构区域，输出标记行和初步区域划分 JSON，
供第二层 AI 做最终裁定。

CLI:
    python table_structure_detector.py input.xlsx [--sheet Sheet1] [--all-sheets]

API:
    from table_structure_detector import detect_structure
    result = detect_structure("input.xlsx", sheet_name="Sheet1")
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))
from detectors import ALL_DETECTORS
from detectors.base import FlaggedRow, EffectiveRange, compute_effective_range


def detect_structure(
    file_path: str,
    sheet_name: str | None = None,
    detectors: list | None = None,
) -> dict:
    """对单个 sheet 执行结构检测。

    Args:
        file_path: Excel 文件路径
        sheet_name: 指定 sheet 名，None 则使用活跃 sheet
        detectors: 自定义检测器类列表，None 则使用全部

    Returns:
        包含 flagged_rows 和 preliminary_regions 的结果字典
    """
    wb = load_workbook(file_path, read_only=False, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        sheet = wb[sheet_name]
    else:
        sheet = wb.active
        sheet_name = sheet.title

    eff = compute_effective_range(sheet)
    sheet_meta = _collect_sheet_meta(sheet, eff)

    detector_classes = detectors or ALL_DETECTORS
    all_flags: list[FlaggedRow] = []
    for cls in detector_classes:
        detector = cls()
        flags = detector.scan(sheet, eff)
        all_flags.extend(flags)

    wb.close()

    grouped = _group_flags_by_row(all_flags)
    regions = _infer_regions(grouped, eff)

    return {
        "file": str(file_path),
        "sheet": sheet_name,
        "total_rows": eff.max_row - eff.min_row + 1,
        "total_cols": eff.col_count,
        "effective_col_range": [eff.min_col, eff.max_col],
        **sheet_meta,
        "flagged_rows": [
            {
                "row": row,
                "flags": [
                    {"detector": f.detector, "detail": f.detail, "severity": f.severity}
                    for f in flags
                ],
            }
            for row, flags in sorted(grouped.items())
        ],
        "preliminary_regions": regions,
    }


def detect_all_sheets(file_path: str) -> list[dict]:
    """对文件中所有 sheet 执行检测"""
    wb = load_workbook(file_path, read_only=False, data_only=True)
    names = wb.sheetnames
    wb.close()

    results = []
    for name in names:
        try:
            result = detect_structure(file_path, sheet_name=name)
            results.append(result)
        except Exception as e:
            results.append({"sheet": name, "error": str(e)})
    return results


def _collect_sheet_meta(sheet, eff: EffectiveRange) -> dict:
    """收集 sheet 级别的辅助元信息"""
    meta = {}

    if sheet.freeze_panes:
        meta["freeze_pane"] = str(sheet.freeze_panes)

    try:
        if sheet.print_area:
            meta["print_area"] = str(sheet.print_area)
    except Exception:
        pass

    hidden_rows = []
    outlined_rows = []
    for r in range(eff.min_row, eff.max_row + 1):
        rd = sheet.row_dimensions.get(r)
        if rd:
            if rd.hidden:
                hidden_rows.append(r)
            if rd.outline_level and rd.outline_level > 0:
                outlined_rows.append({"row": r, "level": rd.outline_level})

    if hidden_rows:
        meta["hidden_rows"] = hidden_rows
    if outlined_rows:
        meta["outlined_rows"] = outlined_rows

    return meta


def _group_flags_by_row(flags: list[FlaggedRow]) -> dict[int, list[FlaggedRow]]:
    grouped: dict[int, list[FlaggedRow]] = {}
    for f in flags:
        grouped.setdefault(f.row, []).append(f)
    return grouped


def _infer_regions(
    grouped: dict[int, list[FlaggedRow]],
    eff: EffectiveRange,
) -> list[dict]:
    """根据标记行推断初步区域划分。

    策略：
    1. 空行 → 子表分隔符
    2. 分隔符之间的连续区域按标记密度判定类型
    3. 首部高密度标记行 → header 候选
    4. 中间无标记行 → data 候选
    5. 尾部标记行 → footer/subtotal 候选
    """
    flagged_set = set(grouped.keys())

    separators = set()
    for row, flags in grouped.items():
        if any(f.detector == "empty_row" for f in flags):
            separators.add(row)

    segments = _split_by_separators(eff.min_row, eff.max_row, separators)

    regions: list[dict] = []
    for sep_row in sorted(separators):
        regions.append({
            "type": "separator",
            "rows": [sep_row, sep_row],
            "confidence": "high",
        })

    for seg_start, seg_end in segments:
        sub_regions = _classify_segment(seg_start, seg_end, grouped, flagged_set)
        regions.extend(sub_regions)

    regions.sort(key=lambda r: r["rows"][0])
    return regions


def _split_by_separators(
    min_row: int, max_row: int, separators: set[int]
) -> list[tuple[int, int]]:
    """按分隔行将行范围分割成多个段"""
    segments = []
    cur_start = min_row

    for r in range(min_row, max_row + 1):
        if r in separators:
            if cur_start < r:
                segments.append((cur_start, r - 1))
            cur_start = r + 1

    if cur_start <= max_row:
        segments.append((cur_start, max_row))

    return segments


def _classify_segment(
    start: int, end: int,
    grouped: dict[int, list[FlaggedRow]],
    flagged_set: set[int],
) -> list[dict]:
    """对一个连续段进行区域分类"""
    if start > end:
        return []

    total = end - start + 1
    if total <= 0:
        return []

    header_end = start - 1
    for r in range(start, end + 1):
        if r in flagged_set:
            flags = grouped[r]
            is_header_signal = any(
                f.detector in ("merged_cells", "format_change", "type_mutation", "column_count")
                for f in flags
            )
            if is_header_signal:
                header_end = r
            else:
                break
        else:
            break

    footer_start = end + 1
    for r in range(end, start, -1):
        if r in flagged_set:
            flags = grouped[r]
            is_footer_signal = any(
                f.detector in ("keyword_match", "numeric_stats")
                for f in flags
            )
            if is_footer_signal:
                footer_start = r
            else:
                break
        else:
            break

    regions = []

    if header_end >= start:
        meta_end = start - 1
        for r in range(start, header_end + 1):
            if r in grouped:
                flags = grouped[r]
                is_meta = any(
                    f.detector == "keyword_match" and "meta" in f.detail
                    for f in flags
                ) or any(
                    f.detector == "column_count" and "单值行" in f.detail
                    for f in flags
                )
                if is_meta:
                    meta_end = r
                else:
                    break
            else:
                break

        if meta_end >= start:
            regions.append({
                "type": "meta_info",
                "rows": [start, meta_end],
                "confidence": "medium",
            })
            if meta_end < header_end:
                regions.append({
                    "type": "header",
                    "rows": [meta_end + 1, header_end],
                    "confidence": "high",
                })
        else:
            regions.append({
                "type": "header",
                "rows": [start, header_end],
                "confidence": "high",
            })

    data_start = max(start, header_end + 1)
    data_end = min(end, footer_start - 1)

    if data_start <= data_end:
        interior_flags = [
            r for r in range(data_start, data_end + 1) if r in flagged_set
        ]

        if interior_flags:
            sub_splits = _find_interior_splits(
                data_start, data_end, grouped, interior_flags
            )
            regions.extend(sub_splits)
        else:
            regions.append({
                "type": "data",
                "rows": [data_start, data_end],
                "confidence": "high",
            })

    if footer_start <= end:
        for r in range(footer_start, end + 1):
            if r not in flagged_set:
                continue
            flags = grouped[r]
            is_subtotal = any(
                f.detector in ("numeric_stats", "keyword_match")
                and any(kw in f.detail for kw in ("求和", "合计", "小计", "总计", "summary"))
                for f in flags
            )
            if is_subtotal:
                regions.append({
                    "type": "subtotal",
                    "rows": [r, r],
                    "confidence": "high",
                })
            else:
                regions.append({
                    "type": "footer",
                    "rows": [r, r],
                    "confidence": "medium",
                })

    return regions


def _find_interior_splits(
    start: int, end: int,
    grouped: dict[int, list[FlaggedRow]],
    flagged_rows: list[int],
) -> list[dict]:
    """处理数据区内部的标记行（如穿插的小计行、子表头）"""
    regions = []
    data_start = start

    for fr in sorted(flagged_rows):
        flags = grouped[fr]

        is_subtotal = any(
            f.detector in ("numeric_stats", "keyword_match")
            for f in flags
        )
        is_new_header = any(
            f.detector == "sequence_break" and "重置" in f.detail
            for f in flags
        )

        if is_new_header:
            if data_start < fr:
                regions.append({
                    "type": "data",
                    "rows": [data_start, fr - 1],
                    "confidence": "medium",
                })
            regions.append({
                "type": "header",
                "rows": [fr, fr],
                "confidence": "medium",
            })
            data_start = fr + 1
        elif is_subtotal:
            if data_start < fr:
                regions.append({
                    "type": "data",
                    "rows": [data_start, fr - 1],
                    "confidence": "medium",
                })
            regions.append({
                "type": "subtotal",
                "rows": [fr, fr],
                "confidence": "medium",
            })
            data_start = fr + 1

    if data_start <= end:
        regions.append({
            "type": "data",
            "rows": [data_start, end],
            "confidence": "medium",
        })

    return regions


def main():
    parser = argparse.ArgumentParser(
        description="Excel 表结构检测器 — 识别表头/数据区/表尾"
    )
    parser.add_argument("file", help="Excel 文件路径 (.xlsx)")
    parser.add_argument("--sheet", "-s", default=None, help="指定 sheet 名称")
    parser.add_argument("--all-sheets", "-a", action="store_true", help="扫描所有 sheet")
    parser.add_argument("--indent", type=int, default=2, help="JSON 缩进")

    args = parser.parse_args()

    if not Path(args.file).exists():
        print(f"Error: File not found: {args.file}", file=sys.stderr)
        sys.exit(1)

    if args.all_sheets:
        result = detect_all_sheets(args.file)
    else:
        result = detect_structure(args.file, sheet_name=args.sheet)

    print(json.dumps(result, ensure_ascii=False, indent=args.indent))


if __name__ == "__main__":
    main()
